"""從本機 topic/ 資料匯入司律一試題庫（取代線上爬蟲）。

來源：
  topic/law.exam-main/100..111/  —— 考選部 txt（PUA 標記），無詳解
  topic/114_law_exam.xlsx        —— 114 年，含詳解
  topic/TaiLexi AI - 2024司律國考一試.xlsx —— 113 年，含詳解

執行：python import_local.py
"""
import re
import sys
from pathlib import Path

import db
from scrapers import lagendre

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

TOPIC = Path(__file__).parent / 'topic'


# ─────────────────────────── 100–111 年 txt ───────────────────────────

def parse_local_filename(fname: str) -> tuple[int, str | None, str] | None:
    """支援兩種命名：
       100司-綜合法學(...).txt        → (100, '司', '綜合法學(...)')
       103綜合法學(...).txt           → (103, None, '綜合法學(...)')  # 103 起司律合併
    """
    if 'ANS' in fname or not fname.endswith('.txt'):
        return None
    m = re.match(r'^(\d+)(司|律)-(.+)\.txt$', fname)
    if m:
        return int(m.group(1)), m.group(2), m.group(3)
    m = re.match(r'^(\d+)(綜合法學.+)\.txt$', fname)
    if m:
        return int(m.group(1)), None, m.group(2)
    return None


def import_txt_years() -> list[dict]:
    records: list[dict] = []
    root = TOPIC / 'law.exam-main'
    if not root.is_dir():
        print(f'  ✗ 找不到 {root}')
        return records
    for ydir in sorted(root.iterdir()):
        if not ydir.is_dir() or not ydir.name.isdigit():
            continue
        files = {f.name for f in ydir.iterdir()}
        year_total = 0
        for fname in sorted(files):
            info = parse_local_filename(fname)
            if not info:
                continue
            year, track, subject = info
            ans_name = fname[:-4] + 'ANS.txt'
            if ans_name not in files:
                continue
            qtext = (ydir / fname).read_text(encoding='utf-8')
            atext = (ydir / ans_name).read_text(encoding='utf-8')
            questions = lagendre.parse_questions(qtext)
            answers   = lagendre.parse_answers(atext)
            for q in questions:
                records.append({
                    'exam_type': 'silu', 'year': year, 'subject': subject,
                    'track': track, 'question_no': q['no'], 'question_text': q['text'],
                    'opt_a': q['A'], 'opt_b': q['B'], 'opt_c': q['C'], 'opt_d': q['D'],
                    'answer': answers.get(q['no']), 'explanation': None,
                })
            year_total += len(questions)
        if year_total:
            print(f'  ✓ {ydir.name} 年：{year_total} 題')
    return records


# ─────────────────────────── xlsx 共用：拆題幹/選項 ───────────────────────────

def _find_markers(s: str) -> dict[str, int] | None:
    """在選項文字中依序定位 A→B→C→D 標記位置。
    分隔不一致（換行 / Tab / 直接相連如「…罪B成立…」）都要能抓到。"""
    pos: dict[str, int] = {}
    idx = 0
    for letter in 'ABCD':
        seg = s[idx:]
        # 優先：行首、Tab 或空白後的字母（最可靠的選項標記）
        m = re.search(r'(?:^|[\n\t　 ])' + letter, seg)
        if not m:
            # 退而求其次：中文字/全形標點後緊接的字母（選項相連的情況）
            m = re.search(r'[一-鿿）)】。，、][ 　]?' + letter, seg)
        if not m:
            return None
        pos[letter] = idx + m.end() - 1   # 字母本身的位置
        idx = pos[letter] + 1
    return pos


def split_stem_options(text: str) -> tuple[str, dict[str, str]] | None:
    """把含選項的文字拆成 (題幹, {A,B,C,D})。

    題幹通常以「？」結尾，其後緊接選項 A；以此為主要分界，
    再用 _find_markers 依序切出 4 個選項。"""
    text = text.strip()
    # 題幹/選項分界：「？」後接選項 A
    mb = re.search(r'[？?]\s*(?=A)', text)
    if mb:
        stem = text[:mb.end()].strip()
        opts_text = text[mb.end():]
    else:
        # 無「？A」樣式：退回以行首 A 為界
        m = re.search(r'(?:^|[\n\t])A', text)
        if not m:
            return None
        cut = m.end() - 1
        stem = text[:cut].strip()
        opts_text = text[cut:]

    pos = _find_markers(opts_text)
    if not pos:
        return None
    order = 'ABCD'
    opts: dict[str, str] = {}
    for i, letter in enumerate(order):
        start = pos[letter] + 1   # 跳過字母本身
        end = pos[order[i + 1]] if i < 3 else len(opts_text)
        opts[letter] = re.sub(r'\s+', ' ', opts_text[start:end]).strip()
    stem = re.sub(r'\s+', ' ', stem).strip()
    if not stem or not all(opts.values()):
        return None
    return stem, opts


# ─────────────────────────── 114 年 xlsx ───────────────────────────

def import_114(path: Path) -> list[dict]:
    import openpyxl
    records: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['全']  # 欄：題號 | 科目 | 題目 | 詳解
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row[2]:
            continue
        no, subject, qcell, exp = row[0], row[1], str(row[2]), row[3]
        m = re.match(r'^答案\s*([A-D]+)\s*[:：]\s*(.*)', qcell, re.DOTALL)
        if not m:
            continue
        answer = m.group(1)
        parsed = split_stem_options(m.group(2))
        if not parsed:
            continue
        stem, opts = parsed
        records.append({
            'exam_type': 'silu', 'year': 114, 'subject': str(subject), 'track': None,
            'question_no': int(no), 'question_text': stem,
            'opt_a': opts['A'], 'opt_b': opts['B'], 'opt_c': opts['C'], 'opt_d': opts['D'],
            'answer': answer, 'explanation': str(exp).strip() if exp else None,
        })
    wb.close()
    print(f'  ✓ 114 年：{len(records)} 題（含詳解）')
    return records


# ─────────────────────────── 113 年 xlsx ───────────────────────────

def import_113(path: Path) -> list[dict]:
    import openpyxl
    records: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        if sn in ('統計', 'All'):
            continue
        ws = wb[sn]  # 欄：題目 | AI回答(詳解) | AI解答 | 正確答案 | ...
        n = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]:
                continue
            qcell, exp, answer = str(row[0]), row[1], row[3]
            mno = re.match(r'^\s*(\d+)', qcell)
            no = int(mno.group(1)) if mno else i
            txt = re.sub(r'^\s*\d+\s*\t?', '', qcell, count=1)
            parsed = split_stem_options(txt)
            if not parsed:
                continue
            stem, opts = parsed
            ans = str(answer).strip().upper() if answer else None
            if ans and not re.fullmatch(r'[A-D]+', ans):
                ans = None
            records.append({
                'exam_type': 'silu', 'year': 113, 'subject': sn, 'track': None,
                'question_no': no, 'question_text': stem,
                'opt_a': opts['A'], 'opt_b': opts['B'], 'opt_c': opts['C'], 'opt_d': opts['D'],
                'answer': ans, 'explanation': str(exp).strip() if exp else None,
            })
            n += 1
        if n:
            print(f'  ✓ 113 年 [{sn}]：{n} 題（含詳解）')
    wb.close()
    return records


# ─────────────────────────── 詳解 txt（topic/詳解/<年>/） ───────────────────────────

def parse_explanation_file(text: str) -> dict[int, str]:
    """把詳解檔拆成 {題號: 詳解}。
    每題以「【第N題】」分隔；詳解取「答：」起的內容（去掉重述的題目與選項）。"""
    out: dict[int, str] = {}
    parts = re.split(r'【第\s*(\d+)\s*題】', text)
    # parts = [前言, '1', body1, '2', body2, ...]
    for i in range(1, len(parts), 2):
        no = int(parts[i])
        body = parts[i + 1] if i + 1 < len(parts) else ''
        m = re.search(r'答\s*[：:]', body)
        expl = (body[m.start():] if m else body).strip()
        if expl:
            out[no] = expl
    return out


def explanation_filename_meta(fname: str) -> tuple[int, str] | None:
    """'110綜合法學(一)(刑法…)_詳解.txt' → (110, '綜合法學(一)(刑法…)')"""
    m = re.match(r'^(\d+)(.+?)_詳解\.txt$', fname)
    if not m:
        return None
    return int(m.group(1)), m.group(2)


def import_explanations() -> int:
    root = TOPIC / '詳解'
    if not root.is_dir():
        print('  （無 topic/詳解/ 目錄，略過）')
        return 0
    total = 0
    for ydir in sorted(root.iterdir()):
        if not ydir.is_dir():
            continue
        for f in sorted(ydir.glob('*.txt')):
            meta = explanation_filename_meta(f.name)
            if not meta:
                continue
            year, subject = meta
            expls = parse_explanation_file(f.read_text(encoding='utf-8'))
            applied = sum(
                db.set_explanation_by_key('silu', year, subject, no, expl)
                for no, expl in expls.items()
            )
            print(f'  ✓ {year} {subject}：套用 {applied}/{len(expls)} 題')
            total += applied
    return total


# ─────────────────────────── 主流程 ───────────────────────────

def main():
    print('=== 從本機 topic/ 匯入司律一試題庫 ===\n')
    db.init_db()

    print('【1/3】100–111 年（考選部 txt）')
    recs = import_txt_years()

    print('\n【2/3】114 年（xlsx，含詳解）')
    recs += import_114(TOPIC / '114_law_exam.xlsx')

    print('\n【3/3】113 年（xlsx，含詳解）')
    recs += import_113(TOPIC / 'TaiLexi AI - 2024司律國考一試.xlsx')

    print('\n【寫入資料庫】')
    inserted = db.insert_questions(recs)
    have_ans = sum(1 for r in recs if r['answer'])
    have_exp = sum(1 for r in recs if r['explanation'])
    print(f'  解析題數：{len(recs)}，新寫入：{inserted}')
    print(f'  有答案：{have_ans}（{round(have_ans / len(recs) * 100)}%）')
    print(f'  內嵌詳解（xlsx）：{have_exp}')

    print('\n【套用詳解 txt（topic/詳解/）】')
    applied = import_explanations()
    print(f'  共套用 {applied} 題詳解')
    years = db.get_filters()['silu']['years']
    print(f'  題庫年份：{years}')
    print('\n✓ 完成，執行 python app.py 開始刷題')


if __name__ == '__main__':
    main()
